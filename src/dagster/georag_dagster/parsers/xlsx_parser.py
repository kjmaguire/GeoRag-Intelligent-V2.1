"""Excel Parser — handles both .xlsx/.xlsm (via Polars/openpyxl) and .xls (via xlrd).

Accepts an Excel file path (.xlsx, .xlsm, or .xls), a sheet name, and a sheet
type.  The file is loaded via the appropriate backend:

  .xlsx / .xlsm — polars.read_excel (openpyxl / fastexcel backend)
  .xls          — xlrd 1.x (the only free library that reads the legacy BIFF OLE
                  format; xlrd >= 2.0 explicitly dropped .xls support)

After loading, both paths export to an in-memory CSV buffer and delegate to the
same CSV parser for the given sheet_type, reusing all alias-matching, validation,
and quality-metric logic without duplication.

Legacy-format behaviours:
  - Merged cells are detected (xlrd exposes sheet.merged_cells).  Sprint 4 does
    NOT auto-unmerge; a structured warning is emitted instead.
  - Multi-row headers: if row 0 is mostly empty but row 1 has full coverage, a
    warning is emitted.  Row 0 is still used as the header (no auto-switch).
  - Formula cells: xlrd returns cached values only (live formulas unavailable).
    This is logged at debug level and is not an error.
  - xls_legacy_format_detected info warning is emitted for all .xls files.

Parse quality metrics are emitted as structured log output so the caller can
record them in Dagster materialisation metadata.

NOTE: Do NOT add `from __future__ import annotations` to this file.
Dagster 1.13 Config classes use Pydantic for type introspection and that import
breaks runtime annotation evaluation.
"""

import hashlib
import io
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

import polars as pl

logger = logging.getLogger(__name__)

PARSER_VERSION = "1.2.0"  # 2026-05-23 — added enumerate_sheets for multi-sheet auto-dispatch

# Supported sheet types map directly to the existing CSV parsers.
SheetType = Literal["collar", "survey", "lithology", "sample"]

# Extension sets for routing to the correct read backend.
_XLSX_EXTS = frozenset({".xlsx", ".xlsm"})
_XLS_EXTS = frozenset({".xls"})


# ---------------------------------------------------------------------------
# Sheet enumeration (added 2026-05-23 per XLSX audit gap #1)
# ---------------------------------------------------------------------------

@dataclass
class SheetMeta:
    """Per-sheet metadata returned by :func:`enumerate_sheets`.

    Used by ``silver_xlsx`` to fan out a multi-sheet workbook to the
    matching CSV parser per sheet, instead of silently dropping
    everything past Sheet 1.
    """

    name: str                  # sheet name as it appears in the workbook
    headers: list[str]         # first-row column names (may be empty)
    row_count: int             # data rows below the header (0 = empty sheet)
    sheet_type: str            # collar | survey | lithology | sample | unknown
    classify_confidence: float # 0.0-1.0 from the header classifier
    hidden: bool               # True if the sheet is hidden / very_hidden


def enumerate_sheets(path: str) -> list[SheetMeta]:
    """Return one :class:`SheetMeta` per sheet in the workbook.

    Hidden sheets are reported with ``hidden=True`` so the caller can
    choose to skip them (the typical correct behaviour — hidden sheets
    in an industry template are usually scratchpads / lookup tables
    that aren't data).

    Both modern (.xlsx/.xlsm) and legacy (.xls) workbooks are handled
    via their respective backends. Sheet classification reuses the
    project's :func:`_sheet_classifier.classify_sheet_type` so the same
    header → type rules apply everywhere.
    """
    # Deferred import — keeps the parser module lightweight at load.
    from georag_dagster.parsers._sheet_classifier import classify_sheet_type  # noqa: PLC0415

    ext = Path(path).suffix.lower()
    out: list[SheetMeta] = []

    if ext in _XLSX_EXTS:
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError as exc:  # pragma: no cover — openpyxl is in the image
            raise RuntimeError(
                "openpyxl unavailable — required for XLSX sheet enumeration"
            ) from exc

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # openpyxl sheet_state: 'visible' | 'hidden' | 'veryHidden'
                hidden = (getattr(ws, "sheet_state", "visible") or "visible") != "visible"
                # Pull the first row as headers; bail out if the sheet is empty.
                headers: list[str] = []
                row_count = 0
                first_row_consumed = False
                for row in ws.iter_rows(values_only=True):
                    if not first_row_consumed:
                        headers = [
                            (str(c) if c is not None else "") for c in row
                        ]
                        first_row_consumed = True
                        continue
                    if any(c is not None and str(c).strip() for c in row):
                        row_count += 1
                sheet_type, confidence = classify_sheet_type(headers)
                out.append(SheetMeta(
                    name=sheet_name,
                    headers=headers,
                    row_count=row_count,
                    sheet_type=sheet_type,
                    classify_confidence=confidence,
                    hidden=hidden,
                ))
        finally:
            wb.close()
        return out

    if ext in _XLS_EXTS:
        try:
            import xlrd  # noqa: PLC0415
        except ImportError as exc:
            raise RuntimeError(
                "xlrd unavailable — required for legacy .xls enumeration"
            ) from exc

        wb = xlrd.open_workbook(path, formatting_info=True)
        for sheet_name in wb.sheet_names():
            sheet = wb.sheet_by_name(sheet_name)
            # xlrd visibility: 0 = visible, 1 = hidden, 2 = very_hidden
            hidden = getattr(sheet, "visibility", 0) != 0
            headers: list[str] = []
            row_count = 0
            if sheet.nrows > 0:
                headers = [
                    (str(v) if v is not None else "")
                    for v in sheet.row_values(0)
                ]
                row_count = max(0, sheet.nrows - 1)
            sheet_type, confidence = classify_sheet_type(headers)
            out.append(SheetMeta(
                name=sheet_name,
                headers=headers,
                row_count=row_count,
                sheet_type=sheet_type,
                classify_confidence=confidence,
                hidden=hidden,
            ))
        return out

    raise ValueError(
        f"xlsx_parser.enumerate_sheets: unsupported extension {ext!r} for {path!r}"
    )


# ---------------------------------------------------------------------------
# Result dataclass
# ---------------------------------------------------------------------------

@dataclass
class ExcelParseResult:
    """Container for a completed Excel sheet parse run.

    Mirrors the shape of the underlying CSV parse result so that the Silver
    asset can handle all four sheet types uniformly.
    """

    source_file: str
    sheet_name: str
    sheet_type: SheetType
    format: Literal["xlsx", "xls", "xlsm"]  # which backend was used
    total_rows: int
    valid_rows: int
    skipped_rows: int
    parse_quality_pct: float
    unmapped_columns: list
    records: list            # validated row dicts — same structure as CSV parse result
    skipped_details: list    # each entry has 'reason' key (and 'raw', 'row' keys)
    column_map: dict         # canonical → original column name from the CSV parser
    assay_columns: list      # only populated for sample sheets
    warnings: list[dict] = field(default_factory=list)
    provenance: dict[str, Any] = field(default_factory=dict)


# Backward-compatible alias — existing code that imports XlsxParseResult still works.
XlsxParseResult = ExcelParseResult


# ---------------------------------------------------------------------------
# SHA-256 provenance helper
# ---------------------------------------------------------------------------

def _sha256_file(path: str) -> str:
    """Stream-hash the file at *path*, returning the hex digest."""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# XLS reader — xlrd 1.x backend
# ---------------------------------------------------------------------------

def _detect_merged_cells(sheet) -> list[dict]:
    """Return a list of merged-cell range dicts from an xlrd Sheet object.

    xlrd exposes sheet.merged_cells as a list of (r_low, r_high, c_low, c_high)
    tuples where the ranges are [r_low, r_high) × [c_low, c_high).
    """
    ranges = []
    for r_low, r_high, c_low, c_high in (sheet.merged_cells or []):
        ranges.append({
            "row_start": r_low, "row_end": r_high,
            "col_start": c_low, "col_end": c_high,
        })
    return ranges


def _detect_multi_row_header(sheet, ncols: int) -> bool:
    """Return True if row 0 is mostly empty but row 1 has full coverage.

    'Mostly empty' means > 50 % of cells in row 0 are blank/None.
    'Full coverage' means > 80 % of cells in row 1 are non-blank.
    """
    if sheet.nrows < 2 or ncols == 0:
        return False

    row0_values = sheet.row_values(0)
    row1_values = sheet.row_values(1)

    row0_empty = sum(1 for v in row0_values if v == "" or v is None)
    row1_filled = sum(1 for v in row1_values if v != "" and v is not None)

    row0_empty_pct = row0_empty / ncols
    row1_fill_pct = row1_filled / ncols

    return row0_empty_pct > 0.5 and row1_fill_pct > 0.8


def _xls_to_polars_df(path: str, sheet_name: str) -> tuple[pl.DataFrame, str, list[dict]]:
    """Load an .xls workbook via xlrd and return a Polars DataFrame plus metadata.

    Returns (df, resolved_sheet_name, xls_warnings).
    All cell values are converted to strings for downstream CSV parser compatibility
    (matching the behaviour of polars.write_csv / read_csv round-trip used for xlsx).
    """
    import xlrd  # noqa: PLC0415 — optional dep; guarded by file extension check

    xls_warnings: list[dict] = []

    # formatting_info=True is required for sheet.merged_cells to be populated.
    wb = xlrd.open_workbook(path, formatting_info=True)

    # Resolve sheet name
    if sheet_name:
        try:
            sheet = wb.sheet_by_name(sheet_name)
            resolved_name = sheet_name
        except xlrd.XLRDError:
            # Fall back to first sheet
            logger.warning(
                "xlsx_parser: sheet '%s' not found in '%s' — using first sheet",
                sheet_name, path,
            )
            sheet = wb.sheet_by_index(0)
            resolved_name = sheet.name
    else:
        sheet = wb.sheet_by_index(0)
        resolved_name = sheet.name

    ncols = sheet.ncols
    nrows = sheet.nrows

    # Detect merged cells — warn but do NOT auto-unmerge (Sprint 4 scope)
    merged = _detect_merged_cells(sheet)
    if merged:
        xls_warnings.append({
            "code": "merged_cells_detected",
            "message": (
                f"Sheet '{resolved_name}' has {len(merged)} merged cell range(s). "
                "Auto-unmerge is deferred to a future sprint."
            ),
            "context": {"count": len(merged), "ranges": merged},
        })
        logger.warning(
            "xlsx_parser: '%s' sheet '%s' has %d merged cell range(s)",
            path, resolved_name, len(merged),
        )

    # Detect multi-row header
    if _detect_multi_row_header(sheet, ncols):
        xls_warnings.append({
            "code": "multi_row_header_suspected",
            "message": (
                "Row 0 appears mostly empty; row 1 may be the actual header. "
                "Row 0 is used as the header — verify the source file."
            ),
            "context": {"header_candidate_rows": [0, 1]},
        })
        logger.warning(
            "xlsx_parser: '%s' sheet '%s' may have a multi-row header",
            path, resolved_name,
        )

    if nrows == 0 or ncols == 0:
        return pl.DataFrame(), resolved_name, xls_warnings

    # Extract header row
    header = [str(v) if v != "" else f"col_{i}" for i, v in enumerate(sheet.row_values(0))]

    # Extract data rows — xlrd returns cached cell values; formulas show computed result
    logger.debug(
        "xlsx_parser: xlrd reads cached values for formula cells in '%s'", path
    )

    rows_data: list[list[str]] = []
    for ridx in range(1, nrows):
        row_vals = sheet.row_values(ridx)
        rows_data.append([
            "" if (v is None or (isinstance(v, float) and str(v) == "nan")) else str(v)
            for v in row_vals
        ])

    # Build Polars DataFrame from string columns
    if not rows_data:
        df = pl.DataFrame({col: pl.Series(col, [], dtype=pl.Utf8) for col in header})
    else:
        col_data = {header[i]: [row[i] if i < len(row) else "" for row in rows_data]
                    for i in range(ncols)}
        df = pl.DataFrame(col_data, schema={k: pl.Utf8 for k in col_data})

    return df, resolved_name, xls_warnings


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_xlsx_sheet(
    path: str,
    sheet_name: str,
    sheet_type: SheetType,
) -> ExcelParseResult:
    """Parse a single sheet of an Excel file (.xlsx, .xlsm, or .xls) as the given sheet_type.

    Routing:
      .xlsx / .xlsm → polars.read_excel (openpyxl / fastexcel backend)
      .xls          → xlrd 1.x (legacy BIFF OLE; xlrd >= 2 does not support .xls)

    After loading, both paths serialise to an in-memory CSV buffer and delegate
    to the matching CSV parser.  All alias resolution, row-level validation, and
    quality metrics come from the CSV parser — no duplication of that logic here.

    Parameters
    ----------
    path:
        Absolute path to the Excel file (.xlsx, .xlsm, or .xls).
    sheet_name:
        Name of the sheet to load.  Pass an empty string to use the first sheet.
    sheet_type:
        One of "collar", "survey", "lithology", "sample".  Controls which CSV
        parser is invoked.

    Returns
    -------
    ExcelParseResult (aliased as XlsxParseResult for backward compatibility).
        Contains validated records plus quality metrics.
    """
    path_str = str(path)
    ext = Path(path_str).suffix.lower()

    # Provenance — hash before opening
    sha256_hex = _sha256_file(path_str)
    provenance: dict[str, Any] = {
        "source_file_sha256": sha256_hex,
        "parser_name": "xlsx_parser",
        "parser_version": PARSER_VERSION,
        # source_col_map populated below once the CSV parser resolves column aliases
        "source_col_map": None,
    }

    extra_warnings: list[dict] = []

    if ext in _XLS_EXTS:
        # --- Legacy .xls path (xlrd 1.x) ---
        extra_warnings.append({
            "code": "xls_legacy_format_detected",
            "message": (
                "File is in legacy .xls binary OLE format.  "
                "Parsed via xlrd 1.x (cached cell values only, no live formulas)."
            ),
            "context": {"path": path_str},
        })
        logger.info("xlsx_parser: '%s' is .xls — using xlrd legacy path", path_str)

        try:
            df, resolved_sheet_name, xls_warnings = _xls_to_polars_df(
                path_str, sheet_name
            )
        except Exception as exc:
            logger.error(
                "xlsx_parser: failed to load .xls from '%s' (sheet=%r): %s",
                path_str, sheet_name, exc,
            )
            raise

        extra_warnings.extend(xls_warnings)
        file_format: Literal["xlsx", "xls", "xlsm"] = "xls"

    elif ext in _XLSX_EXTS:
        # --- Modern .xlsx / .xlsm path (Polars / openpyxl) ---
        try:
            if sheet_name:
                df = pl.read_excel(path_str, sheet_name=sheet_name)
                resolved_sheet_name = sheet_name
            else:
                df = pl.read_excel(path_str)
                resolved_sheet_name = "Sheet1"
                try:
                    import openpyxl  # noqa: PLC0415
                    wb = openpyxl.load_workbook(path_str, read_only=True, data_only=True)
                    resolved_sheet_name = wb.sheetnames[0]
                    wb.close()
                except Exception:
                    pass  # openpyxl unavailable or unreadable — placeholder is fine
        except Exception as exc:
            logger.error(
                "xlsx_parser: failed to load Excel from '%s' (sheet=%r): %s",
                path_str, sheet_name, exc,
            )
            raise

        file_format = "xlsm" if ext == ".xlsm" else "xlsx"

    else:
        raise ValueError(
            f"xlsx_parser: unsupported file extension '{ext}' for '{path_str}'. "
            f"Expected one of: .xlsx, .xlsm, .xls"
        )

    filename = Path(path_str).name
    total_rows_raw = len(df)
    logger.info(
        "Excel loaded: file='%s' sheet='%s' rows=%d columns=%d format=%s",
        filename, resolved_sheet_name, total_rows_raw, len(df.columns), file_format,
    )

    # --- Export to in-memory CSV buffer ---
    # Polars writes all dtypes as strings in CSV, which is exactly what the CSV
    # parsers expect (they use infer_schema=False and cast themselves).
    csv_buffer = io.StringIO()
    df.write_csv(csv_buffer)
    csv_buffer.seek(0)

    # --- Delegate to the matching CSV parser ---
    if sheet_type == "collar":
        from georag_dagster.parsers.csv_collar import parse_csv_collars  # noqa: PLC0415
        result = parse_csv_collars(csv_buffer)
        assay_columns: list = []
    elif sheet_type == "survey":
        from georag_dagster.parsers.csv_survey import parse_csv_surveys  # noqa: PLC0415
        result = parse_csv_surveys(csv_buffer)
        assay_columns = []
    elif sheet_type == "lithology":
        from georag_dagster.parsers.csv_lithology import parse_csv_lithology  # noqa: PLC0415
        result = parse_csv_lithology(csv_buffer)
        assay_columns = []
    elif sheet_type == "sample":
        from georag_dagster.parsers.csv_sample import parse_csv_samples  # noqa: PLC0415
        result = parse_csv_samples(csv_buffer)
        assay_columns = getattr(result, "assay_columns", [])
    else:
        raise ValueError(f"xlsx_parser: unknown sheet_type '{sheet_type}'")

    # Populate source_col_map now that the CSV parser has resolved column aliases.
    provenance["source_col_map"] = result.column_map or {}

    excel_result = ExcelParseResult(
        source_file=filename,
        sheet_name=resolved_sheet_name,
        sheet_type=sheet_type,
        format=file_format,
        total_rows=result.total_rows,
        valid_rows=result.valid_rows,
        skipped_rows=result.skipped_rows,
        parse_quality_pct=result.parse_quality_pct,
        unmapped_columns=result.unmapped_columns,
        records=result.records,
        skipped_details=getattr(result, "skipped_details", []),
        column_map=result.column_map,
        assay_columns=assay_columns,
        warnings=extra_warnings,
        provenance=provenance,
    )

    logger.info(
        "Excel parse complete — sheet='%s' type=%s format=%s total=%d valid=%d "
        "skipped=%d quality=%.1f%%",
        resolved_sheet_name, sheet_type, file_format,
        excel_result.total_rows, excel_result.valid_rows,
        excel_result.skipped_rows, excel_result.parse_quality_pct,
    )

    return excel_result
