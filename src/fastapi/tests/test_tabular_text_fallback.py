"""A spreadsheet ingest_tabular cannot classify was dropped entirely.

Routing the ZIP's workbooks to ``ingest_tabular`` fixed the drill-data
half: a workbook with Collars / Survey / Lithology / Assay tabs now
becomes typed rows whichever way it was uploaded. But ``ingest_tabular``'s
answer for a sheet it cannot classify was one ``nothing_classified``
warning and nothing else — so a genuinely non-drill spreadsheet (a sample
dispatch log, a QA/QC summary, a historical production table) ended up in
the system in no form at all. Down the old ZIP path it was at least
searchable text.

The warning's advice is unactionable for the exact case that regressed:
"pass sheet_type explicitly if the headers are unusual" cannot be done for
a file inside an archive, because the archive branch deliberately passes
no hint.
"""
from __future__ import annotations

import sys
import types
from pathlib import Path

import pytest

from app.services.ingest.xlsx_ingester import (
    XLSXIngestResult,
    ingest_xlsx_file,
    land_sheets_as_text,
)


class _FakeConn:
    """Records what would have been written."""

    def __init__(self, *, existing_report: str | None = None) -> None:
        self.existing_report = existing_report
        self.report_args: tuple | None = None
        self.passages: list[str] = []

    async def fetchrow(self, sql: str, *args):
        flat = " ".join(sql.split())
        if flat.startswith("SELECT report_id"):
            if self.existing_report:
                return {"report_id": self.existing_report}
            return None
        if "INSERT INTO silver.reports" in flat:
            self.report_args = args
            return {"report_id": "11111111-1111-1111-1111-111111111111"}
        if "INSERT INTO silver.document_passages" in flat:
            self.passages.append(args[2])
            return {"passage_id": "p"}
        raise AssertionError(f"unexpected statement: {flat[:120]}")


def _workbook(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    collars = wb.active
    collars.title = "Collars"
    collars.append(["hole_id", "easting", "northing"])
    collars.append(["DDH-14", 500000, 4500000])

    dispatch = wb.create_sheet("Dispatch Log")
    dispatch.append(["batch", "lab", "shipped"])
    dispatch.append(["B-001", "ALS Vancouver", "2026-01-02"])

    qaqc = wb.create_sheet("QAQC")
    qaqc.append(["standard", "expected_ppm"])
    qaqc.append(["OREAS-101", 512])

    wb.save(path)


_WS = "a0000000-0000-0000-0000-00000000feed"
_PJ = "b1000000-0000-0000-0000-0000000000a0"


class TestOnlyTheUnclassifiedSheetsBecomeText:
    @pytest.mark.asyncio
    async def test_the_named_sheets_are_indexed(self, tmp_path) -> None:
        path = tmp_path / "drill.xlsx"
        _workbook(path)
        conn = _FakeConn()

        result = await ingest_xlsx_file(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
            only_sheets=frozenset({"Dispatch Log", "QAQC"}),
        )

        assert result.sheets_processed == 2
        assert result.passages_inserted == 2
        text = "\n".join(conn.passages)
        assert "ALS Vancouver" in text
        assert "OREAS-101" in text

    @pytest.mark.asyncio
    async def test_the_typed_sheets_are_not_duplicated_as_text(
        self, tmp_path,
    ) -> None:
        """The whole point of scoping.

        Sending the entire workbook would store every drill row a second
        time in text shape, where it competes with the typed row in the
        recall set and answers the same question worse.
        """
        path = tmp_path / "drill.xlsx"
        _workbook(path)
        conn = _FakeConn()

        await ingest_xlsx_file(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
            only_sheets=frozenset({"Dispatch Log", "QAQC"}),
        )

        text = "\n".join(conn.passages)
        assert "DDH-14" not in text
        assert "Collars" not in text

    @pytest.mark.asyncio
    async def test_an_unscoped_call_still_reads_the_whole_workbook(
        self, tmp_path,
    ) -> None:
        """`only_sheets=None` is the pre-existing contract; cluster_runner
        still calls it that way."""
        path = tmp_path / "drill.xlsx"
        _workbook(path)
        conn = _FakeConn()

        result = await ingest_xlsx_file(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
        )

        assert result.sheets_processed == 3

    @pytest.mark.asyncio
    async def test_an_empty_scope_writes_nothing(self, tmp_path) -> None:
        """`frozenset()` is not `None`. Treating them alike would index the
        entire workbook precisely when nothing was asked for."""
        path = tmp_path / "drill.xlsx"
        _workbook(path)
        conn = _FakeConn()

        result = await ingest_xlsx_file(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
            only_sheets=frozenset(),
        )

        assert result.skipped is True
        assert result.skipped_reason == "no_sheets_requested"
        assert conn.passages == []
        assert conn.report_args is None


class TestTheWorkbookHandleIsClosed:
    @pytest.mark.asyncio
    async def test_the_file_can_be_deleted_afterwards(self, tmp_path) -> None:
        """``load_workbook(read_only=True)`` streams from the still-open
        .xlsx zip, so the handle outlives the call unless it is closed.

        On the Hatchet worker that is one leaked descriptor per ingest. On
        Windows it is worse: the enclosing ``TemporaryDirectory`` cannot
        delete the file and the cleanup raises PermissionError over the
        whole run. Found by probing this path, not by reading it.
        """
        path = tmp_path / "drill.xlsx"
        _workbook(path)
        conn = _FakeConn()

        await ingest_xlsx_file(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
        )

        path.unlink()   # raises PermissionError on Windows if still open
        assert not path.exists()

    @pytest.mark.asyncio
    async def test_the_handle_is_closed_even_when_a_sheet_raises(
        self, tmp_path, monkeypatch,
    ) -> None:
        path = tmp_path / "drill.xlsx"
        _workbook(path)

        from app.services.ingest import xlsx_ingester

        def _boom(_ws):
            raise RuntimeError("corrupt sheet")

        monkeypatch.setattr(xlsx_ingester, "_format_sheet_as_text", _boom)

        with pytest.raises(RuntimeError, match="corrupt sheet"):
            await ingest_xlsx_file(
                _FakeConn(), str(path), workspace_id=_WS, project_id=_PJ,
            )

        path.unlink()
        assert not path.exists()


class TestTheReportRowIsHonest:
    @pytest.mark.asyncio
    async def test_the_commodity_is_not_stamped_uranium(
        self, tmp_path,
    ) -> None:
        """Every text-landed spreadsheet carried commodity='uranium', a
        leftover from the Wyoming archive this ingester was written for.

        `silver.reports.commodity` is copied into the Qdrant payload, so a
        gold project's spreadsheets were labelled uranium there. The column
        is nullable and 'unknown' is the truth.
        """
        path = tmp_path / "drill.xlsx"
        _workbook(path)
        conn = _FakeConn()

        await ingest_xlsx_file(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
        )

        assert conn.report_args is not None
        assert "uranium" not in [str(a).lower() for a in conn.report_args]

    @pytest.mark.asyncio
    async def test_the_parser_label_says_how_it_was_read(
        self, tmp_path,
    ) -> None:
        conn = _FakeConn()
        target = tmp_path / "notes.csv"
        target.write_bytes(b"a,b\n1,2\n")

        await land_sheets_as_text(
            conn, path=target,
            sheet_texts=[("notes", "a\tb\n1\t2")],
            total_rows=2,
            workspace_id=_WS, project_id=_PJ,
            parser_used="csv-text",
        )

        assert conn.report_args[-1] == "csv-text"

    @pytest.mark.asyncio
    async def test_an_existing_report_is_reused_not_duplicated(
        self, tmp_path,
    ) -> None:
        path = tmp_path / "drill.xlsx"
        _workbook(path)
        conn = _FakeConn(existing_report="99999999-9999-9999-9999-999999999999")

        result = await ingest_xlsx_file(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
        )

        assert conn.report_args is None, "a second report row was inserted"
        assert result.document_id == "99999999-9999-9999-9999-999999999999"


@pytest.fixture
def csv_io_stub(monkeypatch):
    """georag_geoparsers is a path dependency installed only in the
    container image, and the delimited fallback goes through its
    encoding/delimiter helpers on purpose — these files arrive Latin-1
    from Windows survey software and semicolon-delimited from European
    labs."""
    import io as _io

    mod = types.ModuleType("georag_geoparsers._csv_io")

    def open_csv_with_encoding(path):
        raw = Path(path).read_bytes()
        return _io.StringIO(raw.decode("latin-1")), "latin-1", "sha", len(raw)

    def detect_delimiter(content: str) -> str:
        first = content.splitlines()[0] if content.splitlines() else ""
        return ";" if first.count(";") > first.count(",") else ","

    mod.open_csv_with_encoding = open_csv_with_encoding
    mod.detect_delimiter = detect_delimiter

    pkg = sys.modules.get("georag_geoparsers")
    if pkg is None:
        pkg = types.ModuleType("georag_geoparsers")
        pkg.__path__ = []
        monkeypatch.setitem(sys.modules, "georag_geoparsers", pkg)
    monkeypatch.setitem(sys.modules, "georag_geoparsers._csv_io", mod)
    return mod


class TestADelimitedFileCanAlsoLandAsText:
    @pytest.mark.asyncio
    async def test_an_unclassifiable_csv_becomes_searchable(
        self, tmp_path, csv_io_stub,
    ) -> None:
        from app.services.ingest.xlsx_ingester import ingest_delimited_as_text

        path = tmp_path / "dispatch.csv"
        path.write_bytes(
            b"batch,lab,shipped\nB-001,ALS Vancouver,2026-01-02\n"
        )
        conn = _FakeConn()

        result = await ingest_delimited_as_text(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
        )

        assert result.passages_inserted == 1
        assert "ALS Vancouver" in conn.passages[0]
        assert conn.report_args[-1] == "csv-text"

    @pytest.mark.asyncio
    async def test_a_semicolon_file_is_not_stored_one_column_per_row(
        self, tmp_path, csv_io_stub,
    ) -> None:
        """A European lab's export split on the wrong delimiter would store
        the whole row as a single cell and answer nothing."""
        from app.services.ingest.xlsx_ingester import ingest_delimited_as_text

        path = tmp_path / "eu.csv"
        path.write_bytes(b"batch;lab;shipped\nB-002;ALS Loughrea;2026-02-03\n")
        conn = _FakeConn()

        await ingest_delimited_as_text(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
        )

        assert "B-002\tALS Loughrea\t2026-02-03" in conn.passages[0]

    @pytest.mark.asyncio
    async def test_an_empty_file_writes_nothing(
        self, tmp_path, csv_io_stub,
    ) -> None:
        from app.services.ingest.xlsx_ingester import ingest_delimited_as_text

        path = tmp_path / "empty.csv"
        path.write_bytes(b"\n\n")
        conn = _FakeConn()

        result = await ingest_delimited_as_text(
            conn, str(path), workspace_id=_WS, project_id=_PJ,
        )

        assert result.skipped is True
        assert result.skipped_reason == "empty_file"
        assert conn.report_args is None

    @pytest.mark.asyncio
    async def test_a_missing_file_is_reported_not_raised(
        self, tmp_path, csv_io_stub,
    ) -> None:
        from app.services.ingest.xlsx_ingester import ingest_delimited_as_text

        result = await ingest_delimited_as_text(
            _FakeConn(), str(tmp_path / "nope.csv"),
            workspace_id=_WS, project_id=_PJ,
        )

        assert result.skipped is True
        assert result.skipped_reason == "file_not_found"


class TestTheFallbackReportsWhatItDid:
    """``_land_unclassified_as_text`` must never turn a run that DID write
    typed drill rows into a failure — the typed rows are the valuable
    part."""

    @pytest.mark.asyncio
    async def test_a_successful_index_says_where_the_data_went(
        self, tmp_path,
    ) -> None:
        from app.hatchet_workflows import ingest_tabular as it

        path = tmp_path / "drill.xlsx"
        _workbook(path)

        warning = await it._land_unclassified_as_text(
            _FakeConn(), path=str(path), suffix=".xlsx",
            unclassified=["Dispatch Log", "QAQC"],
            workspace_id=_WS, project_id=_PJ,
        )

        assert warning["code"] == "unclassified_indexed_as_text"
        assert "Dispatch Log" in warning["detail"]
        assert "answerable in chat" in warning["detail"]
        assert "will not appear in the drillhole" in warning["detail"]

    @pytest.mark.asyncio
    async def test_the_passage_count_is_reported_as_a_number_too(
        self, tmp_path,
    ) -> None:
        """The caller adds this to ``rows_written``.

        Counting only typed silver rows made a text-only ingest report
        zero, and IngestionRuns.tsx:79 renders zero as "Finished — no
        data written" — over a run whose own warning says it indexed two
        searchable passages. The count is in the prose as well, but a
        caller re-reading it out of the English is how that number goes
        wrong.
        """
        from app.hatchet_workflows import ingest_tabular as it

        path = tmp_path / "drill.xlsx"
        _workbook(path)

        warning = await it._land_unclassified_as_text(
            _FakeConn(), path=str(path), suffix=".xlsx",
            unclassified=["Dispatch Log", "QAQC"],
            workspace_id=_WS, project_id=_PJ,
        )

        assert warning["passages"] == 2
        assert "2 searchable passage(s)" in warning["detail"]

    @pytest.mark.asyncio
    async def test_a_failed_fallback_is_reported_not_raised(
        self, tmp_path, monkeypatch,
    ) -> None:
        from app.hatchet_workflows import ingest_tabular as it
        from app.services.ingest import xlsx_ingester

        async def _boom(*a, **kw):
            raise RuntimeError("openpyxl exploded")

        monkeypatch.setattr(xlsx_ingester, "ingest_xlsx_file", _boom)

        warning = await it._land_unclassified_as_text(
            _FakeConn(), path=str(tmp_path / "x.xlsx"), suffix=".xlsx",
            unclassified=["Sheet1"],
            workspace_id=_WS, project_id=_PJ,
        )

        assert warning["code"] == "unclassified_not_indexed"
        assert "openpyxl exploded" in warning["detail"]

    @pytest.mark.asyncio
    async def test_a_skipped_fallback_is_reported(
        self, tmp_path, monkeypatch,
    ) -> None:
        from app.hatchet_workflows import ingest_tabular as it
        from app.services.ingest import xlsx_ingester

        async def _skipped(*a, **kw):
            return XLSXIngestResult(
                file_path="x", document_id=None, sheets_processed=0,
                rows_total=0, passages_inserted=0,
                skipped=True, skipped_reason="empty_workbook",
            )

        monkeypatch.setattr(xlsx_ingester, "ingest_xlsx_file", _skipped)

        warning = await it._land_unclassified_as_text(
            _FakeConn(), path=str(tmp_path / "x.xlsx"), suffix=".xlsx",
            unclassified=["Sheet1"],
            workspace_id=_WS, project_id=_PJ,
        )

        assert warning["code"] == "unclassified_not_indexed"
        assert "empty_workbook" in warning["detail"]

    @pytest.mark.asyncio
    async def test_a_long_sheet_list_is_truncated_in_the_message(
        self, tmp_path, monkeypatch,
    ) -> None:
        from app.hatchet_workflows import ingest_tabular as it
        from app.services.ingest import xlsx_ingester

        async def _ok(*a, **kw):
            return XLSXIngestResult(
                file_path="x", document_id="d", sheets_processed=9,
                rows_total=90, passages_inserted=9,
            )

        monkeypatch.setattr(xlsx_ingester, "ingest_xlsx_file", _ok)

        warning = await it._land_unclassified_as_text(
            _FakeConn(), path=str(tmp_path / "x.xlsx"), suffix=".xlsx",
            unclassified=[f"S{i}" for i in range(9)],
            workspace_id=_WS, project_id=_PJ,
        )

        assert "(+4 more)" in warning["detail"]


class TestTheOrphanWarningReachesTheProgressRow:
    """`orphaned_intervals` was appended AFTER mark_completed_by_run had
    already serialised `warnings` into the row, so it reached only the
    Hatchet run output object — which is exactly the failure
    mark_completed_by_run's docstring cites as its reason for existing,
    quoting this warning's own text as the example.
    """

    def test_the_orphan_block_runs_before_the_terminal_write(self) -> None:
        from pathlib import Path as _P

        from app.hatchet_workflows import ingest_tabular as it

        source = _P(it.__file__).read_text(encoding="utf-8")
        orphan_at = source.index('"code": "orphaned_intervals"')
        write_at = source.index("await _progress.mark_completed_by_run(")
        assert orphan_at < write_at, (
            "the orphan warning is computed after the terminal write, so it "
            "never reaches silver.ingest_progress.warnings and the Ingestion "
            "Runs page cannot show it"
        )

    def test_it_is_computed_only_once(self) -> None:
        from pathlib import Path as _P

        from app.hatchet_workflows import ingest_tabular as it

        source = _P(it.__file__).read_text(encoding="utf-8")
        assert source.count('"code": "orphaned_intervals"') == 1
