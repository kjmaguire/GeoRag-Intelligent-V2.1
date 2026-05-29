"""Bronze backfill from a pre-staged dir of xlsx files.

Companion to bronze_backfill_uranium_zip.py — the same per-row INSERT
logic, but the input is a directory tree of xlsx files instead of a
ZIP archive. This is the path that ran for the 2026-05-20 backfill
when the 186 GB outer ZIP couldn't be mounted into the container —
we stream-extracted the 11 xlsx files on the WSL host into
/home/georag/uranium_xlsx_staged, docker-cp'd that dir into the
container, then ran this script against it.

Each xlsx is inspected for its shape (collar / assay / unknown)
by reading the first non-blank row as the header and grepping for
geological keywords. Collar-shaped files go to bronze.raw_collar_entries;
assay-shaped go to bronze.raw_assay_submissions.

Run
---
    docker exec georag-fastapi sh -c "cd /app && python \
      scripts/bronze_backfill_xlsx_dir.py \
        --dir /tmp/uranium_xlsx_staged \
        --workspace-id a0000000-0000-0000-0000-000000000001"
"""
from __future__ import annotations

import argparse
import asyncio
import io
import json
import logging
import os
import sys
import uuid
from typing import Any

import asyncpg
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("bronze_xlsx_dir")


# Hole-id markers across the 2026-05-20 Uranium_Logs_ALL.zip dialects:
#   Hole #, HOLEID, hole_id, drillhole, ddh, borehole
#   Label (used in 2008/2009 State Lease "Summary" tables — the value
#     IS a hole id like GH08-351 / SRE09-003)
_COLLAR_KEYWORDS = ("hole", "drillhole", "ddh", "borehole", "holeid", "label")
_ASSAY_TOKENS = ("au_", "cu_", "g/t", "ppm", "ppb", "_ppm", "_ppb", "_au", "_cu")


def _is_collar_header(header: list[str]) -> bool:
    h = [str(x or "").lower().strip() for x in header]
    has_hole = any(any(k in c for k in _COLLAR_KEYWORDS) for c in h)
    has_coord = any(
        ("east" in c or "utm" in c or "longitude" in c
         or c == "x" or c == "easting" or c == "nrth" or "north" in c)
        for c in h
    )
    return has_hole and has_coord


def _is_assay_header(header: list[str]) -> bool:
    h = [str(x or "").lower() for x in header]
    return any(any(t in c for t in _ASSAY_TOKENS) for c in h)


def _find_col(header: list[str], *needles: str) -> int | None:
    for i, raw in enumerate(header):
        cell = str(raw or "").lower()
        for needle in needles:
            if needle in cell:
                return i
    return None


def _as_num(row: list[Any], col: int | None) -> float | None:
    if col is None or col >= len(row):
        return None
    v = row[col]
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _jsonable(v: Any) -> Any:
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


async def _ensure_source_file(
    conn: asyncpg.Connection, workspace_id: str,
    seaweedfs_key: str, original_filename: str, sha: str,
    size_bytes: int, source_type: str, data_type: str,
) -> str:
    row = await conn.fetchrow(
        """
        INSERT INTO bronze.source_files
          (workspace_id, seaweedfs_key, original_filename, file_sha256,
           file_size_bytes, source_type, data_type)
        VALUES ($1::uuid, $2, $3, $4, $5, $6, $7)
        ON CONFLICT (workspace_id, file_sha256) DO UPDATE
           SET seaweedfs_key = EXCLUDED.seaweedfs_key
        RETURNING id::text
        """,
        workspace_id, seaweedfs_key, original_filename, sha,
        size_bytes, source_type, data_type,
    )
    return row["id"]


async def _ingest_collar_xlsx(
    conn: asyncpg.Connection, workspace_id: str, rows: list[list[Any]],
) -> int:
    if not rows:
        return 0
    header = rows[0]
    # Hole-id column: try the canonical names first, then fall back to
    # 'Hole #' (Shirley/Indian Creek), 'HOLEID' (Delineation),
    # 'Label' (State Lease — value IS a hole code, e.g. GH08-351).
    # NOTE: cannot use `or` chaining — when the match is at index 0,
    # Python sees `0 or next_lookup()` as falsy and falls through.
    hole_col = _find_col(header, "hole id", "ddh", "drillhole id", "hole_id", "borehole id")
    if hole_col is None:
        hole_col = _find_col(header, "hole #", "holeid", "hole#")
    if hole_col is None:
        hole_col = _find_col(header, "label")
    east_col  = _find_col(header, "easting", "east", "utm e", "x_utm")
    north_col = _find_col(header, "northing", "nrth", "north", "utm n", "y_utm")
    elev_col  = _find_col(header, "elev", " rl", "z_utm")
    az_col    = _find_col(header, "azim")
    dip_col   = _find_col(header, "dip")
    depth_col = _find_col(header, "total depth", "td", "depth (m)", "depth_m", "total_depth", "planned td")
    logger.info(
        "  collar cols: hole=%s east=%s north=%s elev=%s td=%s | rows=%d",
        hole_col, east_col, north_col, elev_col, depth_col, len(rows) - 1,
    )
    if hole_col is None or (east_col is None and north_col is None):
        return 0

    inserted = 0
    for r in rows[1:]:
        if not r or (hole_col is not None and len(r) <= hole_col):
            continue
        hid = r[hole_col]
        if hid is None or str(hid).strip() == "":
            continue
        await conn.execute(
            """
            INSERT INTO bronze.raw_collar_entries
              (workspace_id, hole_id, easting, northing, elevation,
               azimuth, dip, total_depth, raw_row)
            VALUES ($1::uuid, $2, $3, $4, $5, $6, $7, $8, $9::jsonb)
            """,
            workspace_id, str(hid).strip(),
            _as_num(r, east_col), _as_num(r, north_col), _as_num(r, elev_col),
            _as_num(r, az_col), _as_num(r, dip_col), _as_num(r, depth_col),
            json.dumps(dict(zip(
                [str(h) for h in header],
                [_jsonable(c) for c in r],
            ))),
        )
        inserted += 1
    return inserted


async def _ingest_assay_xlsx(
    conn: asyncpg.Connection, workspace_id: str,
    source_file_id: str, batch_id: str, rows: list[list[Any]],
) -> int:
    if not rows:
        return 0
    header = rows[0]
    hole_col   = _find_col(header, "hole id", "ddh", "hole_id")
    sample_col = _find_col(header, "sample", "sampid")
    from_col   = _find_col(header, "from")
    to_col     = _find_col(header, "to ")
    element_cols: list[tuple[int, str, str]] = []
    for i, raw in enumerate(header):
        cell = str(raw or "")
        low = cell.lower()
        if "_ppm" in low or "_ppb" in low or "g/t" in low or " ppm" in low:
            element = cell.split("_")[0].split(" ")[0].strip()
            unit = ("ppb" if "ppb" in low else "g/t" if "g/t" in low else "ppm")
            if element and element[0].isalpha():
                element_cols.append((i, element, unit))

    if hole_col is None or not element_cols:
        return 0

    inserted = 0
    for r in rows[1:]:
        if not r or len(r) <= hole_col:
            continue
        hid = r[hole_col]
        if hid is None or str(hid).strip() == "":
            continue
        sample_id = (
            str(r[sample_col]).strip() if sample_col is not None and sample_col < len(r) and r[sample_col]
            else f"AUTO-{uuid.uuid4().hex[:8]}"
        )
        from_d, to_d = _as_num(r, from_col), _as_num(r, to_col)
        for col_idx, element, unit in element_cols:
            value = _as_num(r, col_idx)
            if value is None:
                continue
            await conn.execute(
                """
                INSERT INTO bronze.raw_assay_submissions
                  (workspace_id, source_file_id, seaweedfs_key,
                   sample_id, hole_id, from_depth, to_depth,
                   element, value, unit, raw_row, import_batch_id)
                VALUES ($1::uuid, $2::uuid, $3, $4, $5, $6, $7, $8, $9, $10, $11::jsonb, $12::uuid)
                """,
                workspace_id, source_file_id, "uranium-logs/xlsx-dir-staged",
                sample_id, str(hid).strip(),
                from_d, to_d, element, value, unit,
                json.dumps({
                    "headers": [str(h) for h in header],
                    "raw_cells": [_jsonable(c) for c in r],
                }),
                batch_id,
            )
            inserted += 1
    return inserted


async def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True)
    ap.add_argument("--workspace-id", required=True)
    ap.add_argument("--batch-id", default=str(uuid.uuid4()))
    args = ap.parse_args()

    if not os.path.isdir(args.dir):
        logger.error("not a directory: %s", args.dir)
        return 2

    user = os.environ["POSTGRES_USER"]
    pwd  = os.environ["POSTGRES_PASSWORD"]
    host = os.environ.get("POSTGRES_DIRECT_HOST", "postgresql")
    port = os.environ.get("POSTGRES_DIRECT_PORT", "5432")
    db   = os.environ["POSTGRES_DB"]
    pg = await asyncpg.connect(f"postgres://{user}:{pwd}@{host}:{port}/{db}")

    counts = {
        "xlsx_files_seen": 0,
        "collars_inserted": 0,
        "assays_inserted": 0,
        "shape_unknown": 0,
    }

    try:
        for root, _, files in os.walk(args.dir):
            for fname in files:
                if not fname.lower().endswith(".xlsx"):
                    continue
                path = os.path.join(root, fname)
                size = os.path.getsize(path)
                logger.info("opening %s (%d bytes)", path, size)
                counts["xlsx_files_seen"] += 1
                try:
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                except Exception as exc:
                    logger.warning("openpyxl failed on %s: %s", path, exc)
                    continue

                ws = wb[wb.sheetnames[0]]
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                header_idx = next(
                    (i for i, r in enumerate(rows)
                     if sum(1 for c in r if c not in (None, "")) >= 2),
                    None,
                )
                if header_idx is None:
                    continue
                rows = rows[header_idx:]
                header = rows[0]

                sf_id = await _ensure_source_file(
                    pg, workspace_id=args.workspace_id,
                    seaweedfs_key=f"uranium-logs-xlsx/{os.path.basename(root)}/{fname}",
                    original_filename=fname,
                    sha=uuid.uuid4().hex,
                    size_bytes=size,
                    source_type="xlsx",
                    data_type=(
                        "collar" if _is_collar_header(header)
                        else "assay" if _is_assay_header(header)
                        else "unknown"
                    ),
                )

                if _is_collar_header(header):
                    n = await _ingest_collar_xlsx(pg, args.workspace_id, rows)
                    counts["collars_inserted"] += n
                    logger.info("  COLLAR +%d", n)
                elif _is_assay_header(header):
                    n = await _ingest_assay_xlsx(
                        pg, args.workspace_id, sf_id, args.batch_id, rows,
                    )
                    counts["assays_inserted"] += n
                    logger.info("  ASSAY +%d", n)
                else:
                    counts["shape_unknown"] += 1
                    logger.info("  unknown shape, headers=%s", [str(h) for h in header[:8]])
    finally:
        await pg.close()

    logger.info("done: %s", json.dumps(counts, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
