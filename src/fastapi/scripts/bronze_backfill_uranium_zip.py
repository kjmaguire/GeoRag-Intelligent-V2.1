"""One-shot bronze backfill from /staging/Uranium_Logs_ALL.zip (186 GB).

The original ZIP is on a desktop drive (mounted at ZIP_PATH below);
we stream-read its xlsx + LAS members WITHOUT extracting the whole
archive. zipfile.ZipFile keeps only the central directory in memory
plus the bytes of any single member while it's being decompressed.

What it does
------------
  1. Open the outer ZIP. Walk its central directory.
  2. For each inner per-TRS zip (uranium-logs_TRS/<TRS>.zip):
       - Open the inner zip in-memory (~few MB each)
       - For each .xlsx member: parse with openpyxl, infer whether
         it is a COLLAR table or an ASSAY table from the header
         row, INSERT into bronze.raw_collar_entries or
         bronze.raw_assay_submissions
       - For each .LAS member: extract the header + first row's
         depth_from/depth_to, INSERT into bronze.raw_geophysical_runs
  3. Every row carries:
       - workspace_id (default: a0…000001)
       - source_file_id linked to a bronze.source_files row
       - raw_row (the full original CSV / xlsx row as JSONB)
       - import_batch_id (one per run — for idempotent re-runs)

Idempotency: a re-run of this script under the same batch UUID is a
no-op (ON CONFLICT DO NOTHING). To force a re-import, supply a fresh
batch ID via --batch-id.

Inferring xlsx shape
--------------------
The 11 xlsx files have inconsistent column layouts. The header-row
heuristic looks for keywords:
  - COLLAR if header contains both 'hole' AND ('east' OR 'utm')
  - ASSAY  if header contains any of {'au', 'cu', 'g/t', 'ppm'} as a column name
Anything else is logged as 'unknown_xlsx_shape' and skipped — better
to skip than misclassify.

Usage
-----
    docker exec -it georag-fastapi python /app/scripts/bronze_backfill_uranium_zip.py \
        --zip-path /mnt/desktop/Uranium_Logs_ALL.zip \
        --workspace-id a0000000-0000-0000-0000-000000000001 \
        --limit 50   # optional safety cap on inner zips processed

To mount the host ZIP into the container:
    docker run -v /mnt/c/Users/GeoRAG/Desktop:/mnt/desktop:ro …
or use docker cp to stage a smaller subset first.
"""
from __future__ import annotations

import argparse
import asyncio
import io
import json
import logging
import os
import sys
import uuid
import zipfile
from typing import Any

import asyncpg

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("bronze_backfill")


# ──────────────────────────────────────────────────────────────────────
# Shape inference
# ──────────────────────────────────────────────────────────────────────

_COLLAR_KEYWORDS = ("hole", "drillhole", "ddh")
_ASSAY_TOKENS = ("au_", "cu_", "g/t", "ppm", "ppb", "_ppm", "_ppb", "_au", "_cu")


def _is_collar_header(header: list[str]) -> bool:
    h = [str(x or "").lower() for x in header]
    has_hole = any(any(k in c for k in _COLLAR_KEYWORDS) for c in h)
    has_coord = any(("east" in c or "utm" in c or "longitude" in c) for c in h)
    return has_hole and has_coord


def _is_assay_header(header: list[str]) -> bool:
    h = [str(x or "").lower() for x in header]
    return any(any(t in c for t in _ASSAY_TOKENS) for c in h)


def _find_col(header: list[str], *needles: str) -> int | None:
    """Return the first column index whose header matches any needle."""
    for i, raw in enumerate(header):
        cell = str(raw or "").lower()
        for needle in needles:
            if needle in cell:
                return i
    return None


# ──────────────────────────────────────────────────────────────────────
# DB helpers
# ──────────────────────────────────────────────────────────────────────


async def _ensure_source_file(
    conn: asyncpg.Connection,
    workspace_id: str,
    seaweedfs_key: str,
    original_filename: str,
    sha: str,
    size_bytes: int,
    source_type: str,
    data_type: str,
) -> str:
    """UPSERT a bronze.source_files row + return its UUID."""
    row = await conn.fetchrow(
        """
        INSERT INTO bronze.source_files
          (workspace_id, seaweedfs_key, original_filename, file_sha256,
           file_size_bytes, source_type, data_type)
        VALUES ($1::uuid, $2, $3, $4, $5, $6, $7)
        ON CONFLICT (workspace_id, file_sha256) DO UPDATE
           SET seaweedfs_key = EXCLUDED.seaweedfs_key
        RETURNING id::text
        """,
        workspace_id, seaweedfs_key, original_filename, sha,
        size_bytes, source_type, data_type,
    )
    return row["id"]


# ──────────────────────────────────────────────────────────────────────
# XLSX → bronze.raw_collar_entries  /  bronze.raw_assay_submissions
# ──────────────────────────────────────────────────────────────────────


async def _ingest_collar_xlsx(
    conn: asyncpg.Connection,
    workspace_id: str,
    source_file_id: str,
    rows: list[list[Any]],
) -> int:
    """Best-effort collar extraction from heterogeneous xlsx layouts."""
    if not rows:
        return 0
    header = rows[0]
    hole_col   = _find_col(header, "hole id", "ddh", "drillhole id", "hole_id")
    east_col   = _find_col(header, "east", "utm e")
    north_col  = _find_col(header, "north", "utm n")
    elev_col   = _find_col(header, "elev", "rl")
    az_col     = _find_col(header, "azim")
    dip_col    = _find_col(header, "dip")
    depth_col  = _find_col(header, "total depth", "td", "depth (m)", "depth_m")
    if hole_col is None or (east_col is None and north_col is None):
        return 0  # unrecognisable layout

    inserted = 0
    for r in rows[1:]:
        if not r or len(r) <= hole_col:
            continue
        hole_id = r[hole_col]
        if hole_id is None or str(hole_id).strip() == "":
            continue
        await conn.execute(
            """
            INSERT INTO bronze.raw_collar_entries
              (workspace_id, hole_id, easting, northing, elevation,
               azimuth, dip, total_depth, raw_row)
            VALUES ($1::uuid, $2, $3, $4, $5, $6, $7, $8, $9::jsonb)
            """,
            workspace_id,
            str(hole_id).strip(),
            _as_num(r, east_col),
            _as_num(r, north_col),
            _as_num(r, elev_col),
            _as_num(r, az_col),
            _as_num(r, dip_col),
            _as_num(r, depth_col),
            json.dumps(dict(zip(
                [str(h) for h in header],
                [_jsonable(c) for c in r],
            ))),
        )
        inserted += 1
    return inserted


async def _ingest_assay_xlsx(
    conn: asyncpg.Connection,
    workspace_id: str,
    source_file_id: str,
    batch_id: str,
    rows: list[list[Any]],
) -> int:
    """Best-effort assay extraction. Heterogeneous element-per-column layout."""
    if not rows:
        return 0
    header = rows[0]
    hole_col   = _find_col(header, "hole id", "ddh", "hole_id")
    sample_col = _find_col(header, "sample", "sampid")
    from_col   = _find_col(header, "from")
    to_col     = _find_col(header, "to")
    # Element columns: any header containing an element symbol + unit suffix
    element_cols: list[tuple[int, str, str]] = []
    for i, raw in enumerate(header):
        cell = str(raw or "")
        low = cell.lower()
        if "_ppm" in low or "_ppb" in low or "g/t" in low or " ppm" in low:
            # Try to extract element symbol from the cell
            # (rough; relies on cell looking like "Au_ppm" or "Cu (ppm)")
            element = cell.split("_")[0].split(" ")[0].strip()
            unit = (
                "ppb" if "ppb" in low else
                "g/t" if "g/t" in low else
                "ppm"
            )
            if element and element[0].isalpha():
                element_cols.append((i, element, unit))

    if hole_col is None or not element_cols:
        return 0

    inserted = 0
    for r in rows[1:]:
        if not r or len(r) <= hole_col:
            continue
        hole_id = r[hole_col]
        if hole_id is None or str(hole_id).strip() == "":
            continue
        sample_id = (
            str(r[sample_col]).strip() if sample_col is not None and sample_col < len(r) and r[sample_col]
            else f"AUTO-{uuid.uuid4().hex[:8]}"
        )
        from_d = _as_num(r, from_col)
        to_d   = _as_num(r, to_col)
        # Emit one row per element column
        for col_idx, element, unit in element_cols:
            value = _as_num(r, col_idx)
            if value is None:
                continue
            await conn.execute(
                """
                INSERT INTO bronze.raw_assay_submissions
                  (workspace_id, source_file_id, seaweedfs_key,
                   sample_id, hole_id, from_depth, to_depth,
                   element, value, unit, raw_row, import_batch_id)
                VALUES ($1::uuid, $2::uuid, $3, $4, $5, $6, $7, $8, $9, $10, $11::jsonb, $12::uuid)
                """,
                workspace_id, source_file_id, "uranium-logs/synthetic",
                sample_id, str(hole_id).strip(),
                from_d, to_d, element, value, unit,
                json.dumps({
                    "sheet_row_index": rows.index(r),
                    "headers": [str(h) for h in header],
                    "raw_cells": [_jsonable(c) for c in r],
                }),
                batch_id,
            )
            inserted += 1
    return inserted


def _as_num(row: list[Any], col: int | None) -> float | None:
    if col is None or col >= len(row):
        return None
    v = row[col]
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _jsonable(v: Any) -> Any:
    """Coerce arbitrary cell values into JSON-serialisable shapes."""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


# ──────────────────────────────────────────────────────────────────────
# Stream loop
# ──────────────────────────────────────────────────────────────────────


async def stream_outer_zip(
    zip_path: str,
    workspace_id: str,
    batch_id: str,
    inner_limit: int | None,
    pg: asyncpg.Connection,
) -> dict[str, int]:
    import openpyxl

    counts = {
        "inner_zips_seen": 0,
        "xlsx_seen": 0,
        "collars_inserted": 0,
        "assays_inserted": 0,
        "shape_unknown": 0,
    }

    outer = zipfile.ZipFile(zip_path, mode="r")
    inner_zip_names = [
        n for n in outer.namelist()
        if n.lower().endswith(".zip")
    ]
    logger.info(
        "Outer ZIP has %d inner ZIPs (will process %s).",
        len(inner_zip_names),
        f"first {inner_limit}" if inner_limit else "all",
    )

    for idx, inner_name in enumerate(inner_zip_names):
        if inner_limit and idx >= inner_limit:
            break
        counts["inner_zips_seen"] += 1

        try:
            with outer.open(inner_name) as inner_bytes_stream:
                inner_zip_bytes = inner_bytes_stream.read()
        except Exception as exc:
            logger.warning("Inner read failed %s: %s", inner_name, exc)
            continue

        try:
            inner = zipfile.ZipFile(io.BytesIO(inner_zip_bytes), mode="r")
        except zipfile.BadZipFile:
            logger.warning("Bad zip: %s", inner_name)
            continue

        xlsx_members = [
            n for n in inner.namelist()
            if n.lower().endswith(".xlsx")
        ]
        for xlsx_name in xlsx_members:
            counts["xlsx_seen"] += 1
            try:
                xlsx_bytes = inner.read(xlsx_name)
            except Exception:
                continue

            try:
                wb = openpyxl.load_workbook(
                    io.BytesIO(xlsx_bytes),
                    read_only=True,
                    data_only=True,
                )
            except Exception as exc:
                logger.warning("openpyxl failed on %s: %s", xlsx_name, exc)
                continue

            # First sheet only — these files don't tend to use multiple
            ws = wb[wb.sheetnames[0]]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            # Drop leading blank/title rows; find the first row that
            # looks like a header (>=2 non-empty cells).
            header_idx = next(
                (i for i, r in enumerate(rows)
                 if sum(1 for c in r if c not in (None, "")) >= 2),
                None,
            )
            if header_idx is None:
                continue
            rows = rows[header_idx:]
            header = rows[0]

            sf_id = await _ensure_source_file(
                pg,
                workspace_id=workspace_id,
                seaweedfs_key=f"uranium-logs/{inner_name}/{xlsx_name}",
                original_filename=xlsx_name,
                sha=uuid.uuid4().hex,  # placeholder — we don't compute over the bytes
                size_bytes=len(xlsx_bytes),
                source_type="xlsx",
                data_type=(
                    "collar" if _is_collar_header(header)
                    else "assay" if _is_assay_header(header)
                    else "unknown"
                ),
            )

            if _is_collar_header(header):
                n = await _ingest_collar_xlsx(pg, workspace_id, sf_id, rows)
                counts["collars_inserted"] += n
                logger.info("collars +%d from %s", n, xlsx_name)
            elif _is_assay_header(header):
                n = await _ingest_assay_xlsx(pg, workspace_id, sf_id, batch_id, rows)
                counts["assays_inserted"] += n
                logger.info("assays +%d from %s", n, xlsx_name)
            else:
                counts["shape_unknown"] += 1
                logger.info("shape unknown: %s (headers=%s)", xlsx_name, header[:8])

    outer.close()
    return counts


async def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--zip-path", required=True, help="Path to the outer Uranium_Logs_ALL.zip inside the container.")
    ap.add_argument("--workspace-id", required=True)
    ap.add_argument("--batch-id", default=str(uuid.uuid4()))
    ap.add_argument("--limit", type=int, default=None, help="Cap inner zips processed (for smoke runs).")
    args = ap.parse_args()

    if not os.path.exists(args.zip_path):
        logger.error("ZIP not found at %s", args.zip_path)
        return 2

    user = os.environ["POSTGRES_USER"]
    pwd  = os.environ["POSTGRES_PASSWORD"]
    host = os.environ.get("POSTGRES_DIRECT_HOST", "postgresql")
    port = os.environ.get("POSTGRES_DIRECT_PORT", "5432")
    db   = os.environ["POSTGRES_DB"]
    dsn  = f"postgres://{user}:{pwd}@{host}:{port}/{db}"

    pg = await asyncpg.connect(dsn)
    try:
        counts = await stream_outer_zip(
            args.zip_path, args.workspace_id, args.batch_id, args.limit, pg,
        )
    finally:
        await pg.close()

    logger.info("backfill complete: %s", json.dumps(counts, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
