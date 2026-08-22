"""XLSX ingester for Wyoming uranium drillhole archive.

Doc-phase 179 — Phase B Tier 1.

Reads spreadsheet content via `openpyxl`, lands sheet data as text
chunks in `silver.document_passages` (each sheet → N rows → N chunks).

For the 11 XLSX files in the WSGS archive, content is likely:
  - Collar tables (hole_id, easting, northing, depth)
  - Assay tables (hole_id, depth_from, depth_to, U_pct)
  - Lithology tables (hole_id, depth_from, depth_to, lithology)

Phase B Tier 1 just captures the data as searchable text. Phase B
Tier 2 will route XLSX content through the column-mapping wizard to
normalize into typed silver tables.
"""
from __future__ import annotations

import asyncio
import contextlib
import hashlib
import logging
from dataclasses import dataclass
from pathlib import Path

import asyncpg

log = logging.getLogger("georag.ingest.xlsx")


@dataclass
class XLSXIngestResult:
    file_path: str
    document_id: str | None
    sheets_processed: int
    rows_total: int
    passages_inserted: int
    skipped: bool = False
    skipped_reason: str | None = None


def _format_sheet_as_text(sheet) -> str:
    """Format an openpyxl worksheet as tab-separated text.

    First row treated as header. Subsequent rows joined with newlines.
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""
    lines = []
    for row in rows:
        cells = ["" if v is None else str(v).strip() for v in row]
        if any(c for c in cells):
            lines.append("\t".join(cells))
    return "\n".join(lines)


#: Characters per stored passage. Matches the PDF path's window so a
#: spreadsheet row and a report paragraph are comparable units of retrieval.
_SHEET_PASSAGE_CHARS = 5000


def _sheet_passages(
    sheet_texts: list[tuple[str, str]],
) -> list[tuple[int, str]]:
    """Split each sheet into as many passages as it needs.

    A sheet used to be rendered to one blob and hard-cut at 8,000 characters
    with a "[...truncated]" marker. On a 12,000-row assay workbook that is
    roughly 1.5 MB of text reduced to about 80 rows — 99.3% of the assays
    discarded, while the passage was still stored, embedded and retrievable,
    so chat answered assay questions from the first 80 rows and looked like
    it had the data. `rows_total` went on reporting the full count, so
    nothing in the result said otherwise.

    Splitting on line boundaries keeps rows whole; a row cut in half is a
    row that answers nothing. Every passage repeats the sheet header so a
    chunk retrieved on its own still says which sheet it came from and,
    when there is more than one, which part.
    """
    passages: list[tuple[int, str]] = []
    ordinal = 0

    for sheet_name, text in sheet_texts:
        header = f"[Sheet: {sheet_name}]"
        body = text or ""

        chunks: list[str] = []
        current: list[str] = []
        size = 0
        for line in body.split("\n"):
            # +1 for the newline this line will be rejoined with.
            if current and size + len(line) + 1 > _SHEET_PASSAGE_CHARS:
                chunks.append("\n".join(current))
                current, size = [], 0
            current.append(line)
            size += len(line) + 1
        if current:
            chunks.append("\n".join(current))
        if not chunks:
            chunks = [""]

        total = len(chunks)
        for index, chunk in enumerate(chunks, start=1):
            label = header if total == 1 else f"{header} (part {index} of {total})"
            passages.append((ordinal, f"{label}\n{chunk}"))
            ordinal += 1

    return passages


async def ingest_xlsx_file(
    conn: asyncpg.Connection,
    xlsx_path: str,
    *,
    workspace_id: str,
    project_id: str | None = None,
    only_sheets: frozenset[str] | None = None,
) -> XLSXIngestResult:
    """Ingest one XLSX into silver.reports + silver.document_passages.

    ``only_sheets`` restricts the work to named worksheets. That is
    what ingest_tabular passes when a workbook classified partly: the
    Collars / Survey / Lithology tabs became typed rows, and the two
    tabs that matched nothing come here so they are at least
    answerable in chat. Passing the whole workbook instead would
    duplicate every drill row as a second, text-shaped copy competing
    with the typed one in the recall set.
    """
    from openpyxl import load_workbook

    p = Path(xlsx_path)
    if not p.is_file():
        return XLSXIngestResult(
            file_path=xlsx_path, document_id=None,
            sheets_processed=0, rows_total=0, passages_inserted=0,
            skipped=True, skipped_reason="file_not_found",
        )

    if only_sheets is not None and not only_sheets:
        return XLSXIngestResult(
            file_path=xlsx_path, document_id=None,
            sheets_processed=0, rows_total=0, passages_inserted=0,
            skipped=True, skipped_reason="no_sheets_requested",
        )

    try:
        # Hard rule 2 — openpyxl is sync and a large workbook is seconds of
        # CPU on the caller's event loop, which is a Hatchet worker's
        # heartbeat thread.
        wb = await asyncio.to_thread(load_workbook, p, read_only=True, data_only=True)
    except Exception as e:
        return XLSXIngestResult(
            file_path=xlsx_path, document_id=None,
            sheets_processed=0, rows_total=0, passages_inserted=0,
            skipped=True, skipped_reason=f"openpyxl_failed:{type(e).__name__}",
        )

    # Build a single combined text per sheet.
    #
    # `wb.close()` is not optional under read_only=True: openpyxl
    # streams from the still-open .xlsx zip rather than reading it into
    # memory, so the handle lives until the workbook is collected. On a
    # long-lived Hatchet worker that is one leaked descriptor per
    # ingest; on Windows it is worse than a leak, because the enclosing
    # TemporaryDirectory then cannot delete the file and the cleanup
    # raises PermissionError over the whole run. Found by probing this
    # path, not by reading it.
    sheet_texts: list[tuple[str, str]] = []
    total_rows = 0
    try:
        for ws in wb.worksheets:
            if only_sheets is not None and ws.title not in only_sheets:
                continue
            text = _format_sheet_as_text(ws)
            if not text:
                continue
            row_count = text.count("\n") + 1
            total_rows += row_count
            sheet_texts.append((ws.title, text))
    finally:
        with contextlib.suppress(Exception):
            wb.close()

    if not sheet_texts:
        return XLSXIngestResult(
            file_path=xlsx_path, document_id=None,
            sheets_processed=0, rows_total=0, passages_inserted=0,
            skipped=True, skipped_reason="empty_workbook",
        )

    return await land_sheets_as_text(
        conn,
        path=p,
        sheet_texts=sheet_texts,
        total_rows=total_rows,
        workspace_id=workspace_id,
        project_id=project_id,
        parser_used="openpyxl",
    )


async def land_sheets_as_text(
    conn: asyncpg.Connection,
    *,
    path: Path,
    sheet_texts: list[tuple[str, str]],
    total_rows: int,
    workspace_id: str,
    project_id: str | None,
    parser_used: str,
) -> XLSXIngestResult:
    """Land already-rendered sheet text as a report plus its passages.

    Format-agnostic on purpose: a workbook and a delimited file differ only
    in how their rows are read, and the second caller was going to copy
    seventy lines of report-dedupe and passage-insert to avoid saying so.
    """
    # SHA + dedupe
    p = path
    sha = hashlib.sha256(p.read_bytes()).hexdigest()
    # Scoped to the project. The lookup had no project or workspace
    # predicate, so the same workbook uploaded into a SECOND project found
    # the FIRST project's report and wrote its passages under it — invisible
    # in the project the user actually uploaded to, and attached to one they
    # may not even be a member of. Two teams sharing a standard assay
    # template is not an exotic way for a workbook sha to collide.
    row = await conn.fetchrow(
        "SELECT report_id::text AS report_id FROM silver.reports "
        "WHERE source_file_sha256 = $1 AND project_id = $2::uuid LIMIT 1",
        sha, project_id,
    )
    if row:
        document_id = row["report_id"]
    else:
        row = await conn.fetchrow(
            """
            INSERT INTO silver.reports
                (report_id, project_id, workspace_id, title, commodity,
                 source_file_sha256, is_scanned, parser_used,
                 created_at, updated_at)
            VALUES (gen_random_uuid(), $1::uuid, $2::uuid, $3, NULL,
                    $4, false, $5,
                    NOW(), NOW())
            RETURNING report_id::text AS report_id
            """,
            project_id, workspace_id, p.stem[:500], sha, parser_used,
        )
        document_id = row["report_id"]

    # One or more passages per sheet. Tabular content is not
    # paragraph-chunked — _sheet_passages splits on row boundaries only, so
    # a row is never cut in half.
    inserted = 0
    for ordinal, text_with_header in _sheet_passages(sheet_texts):
        h = hashlib.sha256(text_with_header.encode()).hexdigest()
        try:
            r = await conn.fetchrow(
                """
                INSERT INTO silver.document_passages
                    (passage_id, document_id, workspace_id, revision_number,
                     text, text_hash, ordinal, chunk_kind, created_at, updated_at)
                VALUES (gen_random_uuid(), $1::uuid, $2::uuid, 1, $3, $4, $5,
                        'table', NOW(), NOW())
                ON CONFLICT (document_id, revision_number, text_hash) DO NOTHING
                RETURNING passage_id
                """,
                document_id, workspace_id, text_with_header, h, ordinal,
            )
            if r:
                inserted += 1
        except Exception as e:
            log.warning("xlsx_ingester.passage_insert_failed err=%s", e)

    return XLSXIngestResult(
        file_path=str(p),
        document_id=document_id,
        sheets_processed=len(sheet_texts),
        rows_total=total_rows,
        passages_inserted=inserted,
    )


async def ingest_delimited_as_text(
    conn: asyncpg.Connection,
    csv_path: str,
    *,
    workspace_id: str,
    project_id: str | None = None,
) -> XLSXIngestResult:
    """Land one CSV/TSV as searchable text.

    The fallback for a delimited file whose headers match no drill sheet
    type. Before this the answer was a `nothing_classified` warning telling
    the user to "pass sheet_type explicitly if the headers are unusual" —
    advice they cannot act on for a file that arrived inside a ZIP, since
    the archive branch deliberately passes no hint. The file was simply not
    in the system.

    Reads through the same ``_csv_io`` helpers the parsers use: these
    arrive as Latin-1 from Windows survey software and semicolon-delimited
    from European labs, and splitting on the wrong delimiter would store
    one column per row.
    """
    import csv  # noqa: PLC0415

    from georag_geoparsers._csv_io import (  # noqa: PLC0415
        detect_delimiter,
        open_csv_with_encoding,
    )

    p = Path(csv_path)
    if not p.is_file():
        return XLSXIngestResult(
            file_path=csv_path, document_id=None,
            sheets_processed=0, rows_total=0, passages_inserted=0,
            skipped=True, skipped_reason="file_not_found",
        )

    try:
        stream, _encoding, _sha, _size = await asyncio.to_thread(
            open_csv_with_encoding, csv_path,
        )
        content = await asyncio.to_thread(stream.read)
        delimiter = detect_delimiter(content)
    except Exception as e:
        return XLSXIngestResult(
            file_path=csv_path, document_id=None,
            sheets_processed=0, rows_total=0, passages_inserted=0,
            skipped=True, skipped_reason=f"csv_read_failed:{type(e).__name__}",
        )

    lines: list[str] = []
    for row in csv.reader(content.splitlines(), delimiter=delimiter):
        cells = ["" if v is None else str(v).strip() for v in row]
        if any(cells):
            # Tab-separated to match the workbook path, so a retrieved
            # passage reads the same whichever file it came from.
            lines.append("\t".join(cells))

    if not lines:
        return XLSXIngestResult(
            file_path=csv_path, document_id=None,
            sheets_processed=0, rows_total=0, passages_inserted=0,
            skipped=True, skipped_reason="empty_file",
        )

    return await land_sheets_as_text(
        conn,
        path=p,
        sheet_texts=[(p.stem, "\n".join(lines))],
        total_rows=len(lines),
        workspace_id=workspace_id,
        project_id=project_id,
        parser_used="csv-text",
    )


__all__ = [
    "ingest_xlsx_file",
    "ingest_delimited_as_text",
    "land_sheets_as_text",
    "XLSXIngestResult",
]
